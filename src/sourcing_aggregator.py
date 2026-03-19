"""
Sourcing Data Aggregator

CSVファイルからソーシングデータを読み込み、MCID単位で集約してレポートを生成する。
"""

from dataclasses import dataclass
from typing import List
import pandas as pd


# 定数定義
REQUIRED_COLUMNS = [
    'mcid',
    'total_t30d_gms_BAU',
    'SSHVE2_SourcedFlag'
]


@dataclass
class AggregatedResult:
    """集約結果を保持するデータクラス"""
    
    mcid: str
    total_gms: float
    sourced_gms: float
    sourced_gms_percent: float
    
    def to_dict(self) -> dict:
        """辞書形式に変換"""
        return {
            'MCID': self.mcid,
            'Total GMS': self.total_gms,
            'Sourced GMS': self.sourced_gms,
            'Sourced GMS %': self.sourced_gms_percent
        }


@dataclass
class ValidationResult:
    """バリデーション結果を保持するデータクラス"""
    
    is_valid: bool
    missing_columns: List[str]
    error_messages: List[str]
    total_rows: int


class CSVLoader:
    """CSVファイルの読み込みを担当"""
    
    def load(self, file_path: str) -> pd.DataFrame:
        """
        CSVファイルを読み込みDataFrameを返す
        
        Args:
            file_path: 入力CSVファイルパス
            
        Returns:
            pd.DataFrame: 読み込んだデータ
            
        Raises:
            FileNotFoundError: ファイルが存在しない場合
            UnicodeDecodeError: エンコーディングエラーの場合
        """
        # ファイルの存在確認
        import os
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")
        
        # UTF-8で読み込みを試行
        try:
            df = pd.read_csv(file_path, encoding='utf-8')
            return df
        except UnicodeDecodeError:
            # UTF-8で失敗した場合はcp932を試行
            try:
                df = pd.read_csv(file_path, encoding='cp932')
                return df
            except UnicodeDecodeError as e:
                raise UnicodeDecodeError(
                    e.encoding,
                    e.object,
                    e.start,
                    e.end,
                    f"ファイルのエンコーディングを読み取れません: {file_path}"
                )


class DataValidator:
    """データのバリデーションを担当"""
    
    REQUIRED_COLUMNS = REQUIRED_COLUMNS
    
    def validate(self, df: pd.DataFrame) -> ValidationResult:
        """
        必須カラムの存在とデータの妥当性を検証
        
        Args:
            df: 検証対象のDataFrame
            
        Returns:
            ValidationResult: 検証結果
        """
        missing_columns = []
        error_messages = []
        
        # 必須カラムの存在チェック
        for col in self.REQUIRED_COLUMNS:
            if col not in df.columns:
                missing_columns.append(col)
                error_messages.append(f"必須カラムが不足しています: {col}")
        
        # データ行数チェック
        total_rows = len(df)
        if total_rows == 0:
            error_messages.append("データ行が存在しません")
        
        # バリデーション結果を返す
        is_valid = len(missing_columns) == 0 and total_rows > 0
        
        return ValidationResult(
            is_valid=is_valid,
            missing_columns=missing_columns,
            error_messages=error_messages,
            total_rows=total_rows
        )
    
    def clean(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        データをクリーニング
        - total_t30d_gms_BAUを数値型に変換（変換失敗は0）
        - SSHVE2_SourcedFlagを文字列型に変換
        - mcidの欠損値を除外
        
        Args:
            df: クリーニング対象のDataFrame
            
        Returns:
            pd.DataFrame: クリーニング済みデータ
        """
        # DataFrameのコピーを作成（元データを変更しない）
        cleaned_df = df.copy()
        
        # total_t30d_gms_BAUを数値型に変換（変換失敗は0）
        if 'total_t30d_gms_BAU' in cleaned_df.columns:
            cleaned_df['total_t30d_gms_BAU'] = pd.to_numeric(
                cleaned_df['total_t30d_gms_BAU'], 
                errors='coerce'
            ).fillna(0)
        
        # SSHVE2_SourcedFlagを文字列型に変換
        if 'SSHVE2_SourcedFlag' in cleaned_df.columns:
            cleaned_df['SSHVE2_SourcedFlag'] = cleaned_df['SSHVE2_SourcedFlag'].astype(str)
        
        # mcidの欠損値を除外
        if 'mcid' in cleaned_df.columns:
            cleaned_df = cleaned_df[cleaned_df['mcid'].notna()]
        
        return cleaned_df


class MCIDAggregator:
    """MCID単位でのデータ集約を担当"""
    
    def aggregate(self, df: pd.DataFrame) -> List[AggregatedResult]:
        """
        MCID単位でデータを集約
        
        処理内容:
        1. mcidでグループ化
        2. 各グループで以下を計算:
           - total_gms = sum(total_t30d_gms_BAU)
           - sourced_gms = sum(total_t30d_gms_BAU where SSHVE2_SourcedFlag == 'Y')
           - sourced_gms_percent = (sourced_gms / total_gms) * 100
        
        Args:
            df: クリーニング済みDataFrame
            
        Returns:
            List[AggregatedResult]: 集約結果のリスト
        """
        results = []
        
        # mcidでグループ化
        grouped = df.groupby('mcid')
        
        for mcid, group in grouped:
            # Total GMSを計算
            total_gms = group['total_t30d_gms_BAU'].sum()
            
            # Sourced GMSを計算
            sourced_gms = self._calculate_sourced_gms(group)
            
            # Sourced GMS %を計算
            sourced_gms_percent = self._calculate_percentage(sourced_gms, total_gms)
            
            # 結果を追加
            results.append(AggregatedResult(
                mcid=str(mcid),
                total_gms=total_gms,
                sourced_gms=sourced_gms,
                sourced_gms_percent=sourced_gms_percent
            ))
        
        return results
    
    def _calculate_sourced_gms(self, group: pd.DataFrame) -> float:
        """
        Sourced GMSを計算
        SSHVE2_SourcedFlag == 'Y' の行のtotal_t30d_gms_BAUを合計
        
        Args:
            group: MCIDでグループ化されたDataFrame
            
        Returns:
            float: Sourced GMSの合計
        """
        sourced_rows = group[group['SSHVE2_SourcedFlag'] == 'Y']
        return sourced_rows['total_t30d_gms_BAU'].sum()
    
    def _calculate_percentage(self, sourced_gms: float, total_gms: float) -> float:
        """
        Sourced GMS %を計算
        total_gmsが0の場合は0を返す
        
        Args:
            sourced_gms: Sourced GMSの値
            total_gms: Total GMSの値
            
        Returns:
            float: Sourced GMS %（0-100の範囲）
        """
        if total_gms == 0:
            return 0.0
        return (sourced_gms / total_gms) * 100


class HTMLReportGenerator:
    """HTMLレポートの生成を担当"""

    def generate(
        self,
        results: List[AggregatedResult],
        output_path: str,
        source_file: str
    ) -> None:
        """
        HTMLレポートを生成

        レポート内容:
        - タイトル: "Sourcing Data Aggregation Report"
        - 生成日時
        - ソースファイル名
        - 集約結果テーブル（MCID, Total GMS, Sourced GMS, Sourced GMS %）
        - 数値フォーマット: カンマ区切り、小数点2桁

        Args:
            results: 集約結果
            output_path: 出力HTMLファイルパス
            source_file: ソースCSVファイルパス
        """
        from datetime import datetime
        import os

        # タイムスタンプを生成
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ソースファイル名を取得
        source_filename = os.path.basename(source_file)

        # HTMLテーブルの行を生成
        table_rows = []
        for result in results:
            row = f'''
            <tr>
                <td>{result.mcid}</td>
                <td class="number">{self._format_number(result.total_gms)}</td>
                <td class="number">{self._format_number(result.sourced_gms)}</td>
                <td class="number">{self._format_percentage(result.sourced_gms_percent)}%</td>
            </tr>'''
            table_rows.append(row)

        table_body = '\n'.join(table_rows)

        # HTML全体を構築
        html = f'''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sourcing Data Aggregation Report</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', 'Meiryo', sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            color: white;
            padding: 40px;
        }}
        .header h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
        .header p {{ opacity: 0.8; font-size: 1.1em; }}
        .metadata {{
            background: rgba(255,255,255,0.1);
            border-radius: 10px;
            padding: 20px;
            margin-top: 20px;
        }}
        .metadata-item {{
            margin-bottom: 10px;
        }}
        .metadata-label {{
            font-size: 0.9em;
            opacity: 0.7;
        }}
        .metadata-value {{
            font-size: 1.1em;
            font-weight: bold;
            margin-left: 10px;
        }}
        .content {{ padding: 40px; }}
        .section {{ margin-bottom: 40px; }}
        .section-title {{
            font-size: 1.5em;
            color: #1a1a2e;
            border-left: 5px solid #667eea;
            padding-left: 15px;
            margin-bottom: 20px;
        }}
        .table-container {{
            background: white;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .data-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .data-table thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}
        .data-table th {{
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        .data-table td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}
        .data-table td.number {{
            text-align: right;
        }}
        .data-table tbody tr:hover {{
            background: #f8f9fa;
        }}
        .data-table tbody tr:nth-child(even) {{
            background: #fafafa;
        }}
        .data-table tbody tr:nth-child(even):hover {{
            background: #f0f0f0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Sourcing Data Aggregation Report</h1>
            <p>MCID単位のソーシングデータ集約レポート</p>

            <div class="metadata">
                <div class="metadata-item">
                    <span class="metadata-label">生成日時:</span>
                    <span class="metadata-value">{timestamp}</span>
                </div>
                <div class="metadata-item">
                    <span class="metadata-label">ソースファイル:</span>
                    <span class="metadata-value">{source_filename}</span>
                </div>
            </div>
        </div>

        <div class="content">
            <div class="section">
                <h2 class="section-title">📋 集約結果（{len(results)}件）</h2>

                <div class="table-container">
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>MCID</th>
                                <th>Total GMS</th>
                                <th>Sourced GMS</th>
                                <th>Sourced GMS %</th>
                            </tr>
                        </thead>
                        <tbody>
                            {table_body}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>
</body>
</html>'''

        # HTMLファイルを保存
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

    def _format_number(self, value: float) -> str:
        """
        数値をカンマ区切りでフォーマット

        Args:
            value: フォーマット対象の数値

        Returns:
            str: カンマ区切りの文字列
        """
        return f'{value:,.2f}'

    def _format_percentage(self, value: float) -> str:
        """
        パーセンテージを小数点2桁でフォーマット

        Args:
            value: フォーマット対象のパーセンテージ値

        Returns:
            str: 小数点2桁の文字列
        """
        return f'{value:.2f}'






class ExcelExporter:
    """Excelファイルの出力を担当"""
    
    def export(
        self,
        results: List[AggregatedResult],
        output_path: str
    ) -> None:
        """
        Excelファイルを生成
        
        シート構成:
        - "集約結果": MCID, Total GMS, Sourced GMS, Sourced GMS %
        
        フォーマット:
        - 数値列: カンマ区切り、小数点2桁
        - パーセンテージ列: パーセンテージ形式、小数点2桁
        - ヘッダー: 太字、背景色
        
        Args:
            results: 集約結果
            output_path: 出力Excelファイルパス
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxlライブラリが必要です。pip install openpyxl でインストールしてください。"
            )
        
        # ワークブックを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "SS2_Sourcing_Status_BySeller"
        
        # ヘッダー行を設定
        headers = ['MCID', 'Total GMS', 'Sourced GMS', 'Sourced GMS %']
        ws.append(headers)
        
        # ヘッダーのスタイリング
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # データ行を追加
        for result in results:
            # MCIDを数値として扱う（数値変換可能な場合）
            try:
                mcid_value = float(result.mcid) if result.mcid.replace('.', '', 1).isdigit() else result.mcid
            except (ValueError, AttributeError):
                mcid_value = result.mcid
            
            ws.append([
                mcid_value,
                result.total_gms,
                result.sourced_gms,
                result.sourced_gms_percent
            ])
        
        # 列幅を調整
        ws.column_dimensions['A'].width = 20  # MCID
        ws.column_dimensions['B'].width = 18  # Total GMS
        ws.column_dimensions['C'].width = 18  # Sourced GMS
        ws.column_dimensions['D'].width = 18  # Sourced GMS %
        
        # 数値フォーマットを適用（ヘッダー行を除く）
        for row in range(2, len(results) + 2):
            # Total GMS (B列) - カンマ区切り、小数点2桁
            ws[f'B{row}'].number_format = '#,##0.00'
            
            # Sourced GMS (C列) - カンマ区切り、小数点2桁
            ws[f'C{row}'].number_format = '#,##0.00'
            
            # Sourced GMS % (D列) - パーセンテージ形式、小数点2桁
            ws[f'D{row}'].number_format = '0.00"%"'
            
            # 数値列を右揃え
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            ws[f'D{row}'].alignment = Alignment(horizontal='right')
        
        # Excelファイルを保存
        wb.save(output_path)


class CSVExporter:
    """CSVファイルの出力を担当"""
    
    def export(
        self,
        results: List[AggregatedResult],
        output_path: str
    ) -> None:
        """
        CSVファイルを生成
        
        Args:
            results: 集約結果
            output_path: 出力CSVファイルパス
        """
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # ヘッダー行を書き込み
            writer.writerow(['MCID', 'Total GMS', 'Sourced GMS', 'Sourced GMS %'])
            
            # データ行を書き込み
            for result in results:
                # MCIDを数値として扱う（数値変換可能な場合）
                try:
                    mcid_value = float(result.mcid) if result.mcid.replace('.', '', 1).isdigit() else result.mcid
                except (ValueError, AttributeError):
                    mcid_value = result.mcid
                
                writer.writerow([
                    mcid_value,
                    result.total_gms,
                    result.sourced_gms,
                    result.sourced_gms_percent
                ])
